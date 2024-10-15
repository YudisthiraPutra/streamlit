import os
import io
import glob
import time
import datetime
import numpy as np
import pandas as pd
import openpyxl # type: ignore
from openpyxl import load_workbook, Workbook # type: ignore
from openpyxl.styles import (Border, Side, Alignment, Font,  # type: ignore
                            PatternFill)
from openpyxl.utils import get_column_letter # type: ignore
from openpyxl.worksheet.table import Table, TableStyleInfo # type: ignore
from collections import Counter
from datetime import datetime
import re
import string
from copy import copy
from IPython.display import display, HTML
import streamlit as st # type: ignore
import json

# # VERTICAL CHECK
# # VERTICAL CHECK
# # VERTICAL CHECK

def mapping_dictionary(wb_path):
    wb = openpyxl.load_workbook(wb_path, data_only=True)
    sheet = wb['Dictionary Intra Antar']
    mappings = {}
    for row in sheet.iter_rows(min_row=2, max_col=3, values_only=True):
        key = str(row[0])  # Kolom A berisi key di file dictionary (fixed)
        description = str(row[1])  # Kolom B berisi deskripsi di file dictionary (fixed)
        description_list = [desc.strip() for desc in
                            description.split(
                                ';')]  # Ini untuk membaca deksripsinya menjadi sebuah list jika ada ','
        mappings[key] = description_list
    return mappings

wb_path = "/Users/ferroyudisthira/Desktop/DSTA_DQAD/V&H_Check/Kodifikasi.xlsx"
mappings = mapping_dictionary(wb_path)

def find_key_by_value(mappings, target_value):
    for key, value in mappings.items():
        if value[0] == target_value:
            return key
    return None

# Gunakan raw string notation untuk menghindari masalah escape sequence
sski_folder = r'/Users/ferroyudisthira/Desktop/DSTA_DQAD/V&H_Check/Sumber_Data_Lama/SSKI'

if os.sep == "/":
    nama_sski_folder = sski_folder.split('/')[-1]
    filenames_sski = glob.glob(sski_folder + r"/*.xlsx")
elif os.sep == "\\":
    nama_sski_folder = sski_folder.split('\\')[-1]
    filenames_sski = glob.glob(sski_folder + r"\*.xlsx")
# Raw string notation untuk pola glob
workbooks_sski = {}
# print("filename ", filenames_sski)

for filename in filenames_sski:
    # Skip file sementara Excel yang dimulai dengan ~$ 
    if '~$' in filename:
        continue
    
    wb_sski = openpyxl.load_workbook(filename)
    sheet_names_sski = wb_sski.sheetnames
    # print(sheet_names_sski)
    for sheet_name in sheet_names_sski:
        sheet_sski = wb_sski[sheet_name]
        
        # Hilangkan semua tanda titik dari nama sheet
        sheet_name_clean = sheet_name.replace('.', '')
        
        # Dapatkan indeks sheet menggunakan find_key_by_value
        sheet_index_sski = find_key_by_value(mappings, sheet_name_clean)
        
        # Pastikan indeks sheet adalah string
        if sheet_index_sski is None:
            sheet_index_sski = sheet_name_clean  # Atau tangani sesuai kebutuhan
        else:
            sheet_index_sski = str(sheet_index_sski)
        
        # Buat nama sheet yang sudah dimodifikasi
        modified_sheet_name_sski = f'{nama_sski_folder}{sheet_index_sski}'
        
        # Tambahkan sheet ke dictionary globals
        globals()[modified_sheet_name_sski] = sheet_sski


def indikator(wb_path, mappings):
    wb = openpyxl.load_workbook(wb_path, data_only=True)
    sheet = wb['Kodifikasi Vertical Check']

    kelompok = []
    indikator_asli = []
    indikator_calculated = []

    last_row = sheet.max_row
    four_digit_keys = [key for key in mappings.keys() if (len(key) == 4 or len(key) == 5)]

    for row in sheet.iter_rows(min_row=2):
        # Check if column 2 contains 'SSKI'
        if 'SSKI' not in str(row[1].value):
            continue  # Skip rows where column 2 does not contain 'SSKI'
        else:
            new_indikator_asli = str(row[1].value)
            new_indikator_calculated = str(row[2].value)
            kelompok.append(row[0].value)
            indikator_asli.append(new_indikator_asli)
            indikator_calculated.append(new_indikator_calculated)

    # Print the outputs
    # print("Kelompok:", kelompok)
    # print("Indikator Asli:", indikator_asli)
    # print("Indikator Calculated:", indikator_calculated)

    # Return the results
    return kelompok, indikator_asli, indikator_calculated

mappings = mapping_dictionary(wb_path)
kelompok, indikator_asli, indikator_calculated = indikator(wb_path, mappings)

def dataframe_bag(indikator_bag, mappings):
    queries = re.split(r'[+\-;/]', indikator_bag)
    grup_data = []
    list_data = []
    paths = []
    indikator_name=[]
    for query in queries:
        token = query.split('.')
        paths.append(query)
        data = token[0] + token[1]
        grup_data.append(data)

    if grup_data:
        list_data = [globals().get(name) for name in grup_data if name in globals()]
    
    token = queries[0].split('.')
    
    results = []

    for ws, query in zip(list_data, queries):
        data_frames = []
        token = query.split('.')
        baris_pembukaan = 1
        nama_indikator = []
        last_found_column_idx = 0

        for i, t in enumerate(token[2:]):
            if t in mappings:
                deskripsi_list = mappings[t]
                ditemukan = False

                for deskripsi in deskripsi_list:
                    desk=deskripsi
                    deskripsi = deskripsi.replace('\u2011', '-').replace('\u00A0', ' ').strip().replace(' ', '').lower()
                    
                    for baris in range(baris_pembukaan, ws.max_row + 1):
                        for kolom_idx in range(last_found_column_idx, len(string.ascii_uppercase)):
                            kolom = string.ascii_uppercase[kolom_idx]
                            nilai_sel = ws[f'{kolom}{baris}'].value
                            if nilai_sel is not None:
                                nilai_sel = str(nilai_sel).replace('\u2011', '-').replace('\u00A0', ' ').strip().replace(' ', '').lower()
                                if nilai_sel == '-':
                                    nilai_sel = 0

                            if nilai_sel == deskripsi:
                                baris_pembukaan = baris
                                nama_indikator.append((t, nilai_sel, baris))
                                ditemukan = True
                                last_found_column_idx = kolom_idx
                                break
                        if ditemukan:
                            break
                    if ditemukan:
                        break

        indikator_name.append(desk)
        kolom_kodifikasi = kolom
        kolom_nilai = ord(kolom_kodifikasi.upper()) - ord('A') + 2
        results.append((nama_indikator, baris_pembukaan))
    
    baris_pembukaan_list = [hasil[1] for hasil in results]

    def create_data_frames(ws, baris_data, baris_kolom):
        column_names = [ws.cell(row=baris_kolom, column=col).value for col in range(2, ws.max_column + 1)]
        data = []
        for col in range(2, ws.max_column + 1):
            cell_value = ws.cell(row=baris_data, column=col).value
            if cell_value is None or cell_value == '-' or str(cell_value).strip() == '':
                cell_value = 0
            data.append(cell_value)
        df = pd.DataFrame([data], columns=column_names)
        df.iloc[:, 1:] = df.iloc[:, 1:].apply(pd.to_numeric, errors='coerce')
        return df
    
    data_frames_dict = {}
    
    for idx, (ws, baris_data) in enumerate(zip(list_data, baris_pembukaan_list)):
        for i in range(3, 8):
            df_1 = create_data_frames(ws, baris_data, i)
            df_1 = df_1.drop(df_1.columns[[0]], axis=1)
            if not df_1.empty:
                first_col = df_1.columns[0]
                if any(isinstance(first_col, int) and 1900 <= first_col <= 2100 for first_col in df_1.columns):
                    baris_kolom = i
                    break
        
        df = create_data_frames(ws, baris_data, i)
        df = df.iloc[:, 1:-2]
        # Bersihkan nama kolom
        data_frames_dict[(ws, idx)] = (baris_kolom, df)
        kolom = df.columns.to_list()
    
    data_frames = []

    for (ws, idx), (baris_kolom, df) in data_frames_dict.items():
        df = df.drop(df.columns[[0]], axis=1)
    
        # Get column names from DataFrame
        columns = list(df.columns)
        # print(len(columns))

        # Initialize lists for new column names
        nama_kolom_baru = []
        bulan_list = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        quarters_list = ['Q1', 'Q2', 'Q3', 'Q4']
        current_year = None
        bulan_index = 0
        quarter_index = 0

        i = 0
        # Check for monthly or quarterly data based on token presence
        if '2' in token or '3' in token or '20' in token or '19' in token or '18' in token or '17' in token or '15' in token or '6' in token or '14' in token or '12' in token or '11a' in token or '10' in token or '8' in token or '9' in token or '7' in token:  # Monthly data
            while i < len(columns):
                if columns[i] is None:
                    # Handle None columns
                    if current_year is not None:
                        while i < len(columns) and columns[i] is None:
                            nama_kolom_baru.append(f"{bulan_list[bulan_index]} {current_year}")
                            bulan_index += 1
                            if bulan_index == len(bulan_list):  # Reset index after December
                                current_year += 1
                                bulan_index = 0
                            i += 1
                    continue
                else:
                    if columns[i] == '2023**':
                        current_year = 2023
                    else:
                        current_year = int(columns[i])  
                    if i + 1 < len(columns) and columns[i + 1] is None:
                        nama_kolom_baru.append(f"{bulan_list[bulan_index]} {current_year}")
                        bulan_index += 1
                    else:
                        nama_kolom_baru.append(str(current_year))
                    i += 1
            
            # Assign new column names
            df.columns = nama_kolom_baru
            data_frames.append(df)
            
        elif '4' in token or '16a' in token:  # Quarterly data
            while i < len(columns):
                if columns[i] is None:
                    if current_year is not None:
                        while i < len(columns) and columns[i] is None:
                            nama_kolom_baru.append(f"{quarters_list[quarter_index]} {current_year}")
                            quarter_index += 1
                            if quarter_index == len(quarters_list):  # Reset index after Q4
                                current_year += 1
                                quarter_index = 0
                            i += 1
                    continue
                else:
                    if columns[i] == '2023**':
                        current_year = 2023
                    else:
                        current_year = int(columns[i])  
                    if i + 1 < len(columns) and columns[i + 1] is None:
                        nama_kolom_baru.append(f"{quarters_list[quarter_index]} {current_year}")
                        quarter_index += 1
                    else:
                        nama_kolom_baru.append(str(current_year))
                    i += 1
            
            # Assign new column names
            df.columns = nama_kolom_baru
            data_frames.append(df)

        elif '5a' in token:  # Claims and liabilities
            year_count = len([x for x in columns if x is not None])  # Hitung berapa banyak tahun di dalam kolom
            
            while i < len(columns):
                if columns[i] is not None and columns.count(columns[i]) == 1:  # Case for a single year
                    # Rename the year column to "Claims + year"
                    nama_kolom_baru.append(f"Claims {columns[i]}")
                    current_year = columns[i]
                    i += 1
                    
                    # Rename subsequent None columns
                    while i < len(columns) and columns[i] is None:
                        nama_kolom_baru.append(f"Liabilities {current_year}")
                        i += 1
        
                elif columns[i] is not None and columns.count(columns[i]) == 13:  # Kasus untuk 13 kolom
                    # Pertama, rename tahun seperti di kasus satu tahun
                    nama_kolom_baru.append(f"Claims {columns[i]}")
                    current_year = columns[i]
                    i += 1
        
                    # Rename subsequent None columns
                    while i < len(columns) and columns[i] is None:
                        nama_kolom_baru.append(f"Liabilities {current_year}")
                        i += 1
        
                    # Setelah itu lanjutkan dengan loop untuk bulan-bulan
                    bulan_index = 0  # Reset month index for the new year
        
                    while bulan_index < len(bulan_list) and i < len(columns):
                        nama_kolom_baru.append(f"Claims {bulan_list[bulan_index]} {current_year}")
                        i += 1  # Move to the next column
                        
                        if i < len(columns) and columns[i] is None:
                            nama_kolom_baru.append(f"Liabilities {bulan_list[bulan_index]} {current_year}")
                            i += 1  # Move to the next column
        
                        bulan_index += 1
        
                elif columns[i] is not None and columns.count(columns[i]) >1 and  columns.count(columns[i])<=12:  # Kasus untuk 12 kolom
                    current_year = columns[i]  # Tetapkan current_year
                    i += 1
                    bulan_index = 0
        
                    while bulan_index < len(bulan_list) and i < len(columns):
                        nama_kolom_baru.append(f"Claims {bulan_list[bulan_index]} {current_year}")
                        i += 1  # Move to the next column
        
                        nama_kolom_baru.append(f"Liabilities {bulan_list[bulan_index]} {current_year}")
                        i += 1  # Move to the next column
        
                        bulan_index += 1
        
                else:
                    # Move to the next column if no valid conditions are met
                    i += 1
        
            # print(nama_kolom_baru)
            df.columns = nama_kolom_baru
            data_frames.append(df)

        elif '5b' in token or '5c' in token:  # Claims and liabilities
            subcategories = ['Total', 'Loan & Deposits', 'Debt Securities', 'Other Instruments']
            year_count = len([x for x in columns if x is not None])  # Hitung berapa banyak tahun di dalam kolom
            
            processed_years = []  # List to keep track of processed years

            while i < len(columns):
                if columns[i] is not None and columns.count(columns[i]) == 1 and columns[i] not in processed_years:
                    current_year = columns[i]
                    processed_years.append(current_year)
                    
                    # Loop through the subcategories for Claims
                    for sub in subcategories:
                        nama_kolom_baru.append(f"Claims {sub} {current_year}")
                        i += 1
                    
                    for sub in subcategories:
                        nama_kolom_baru.append(f"Liabilities {sub} {current_year}")
                        i += 1
                
                elif columns[i] is not None and columns.count(columns[i]) == 13 and columns[i] not in processed_years:
                    current_year = columns[i]
                    processed_years.append(current_year)
            
                    # Loop through the subcategories for Claims
                    for sub in subcategories:
                        nama_kolom_baru.append(f"Claims {sub} {current_year}")
                        i += 1
                    
                    for sub in subcategories:
                        nama_kolom_baru.append(f"Liabilities {sub} {current_year}")
                        i += 1
                    
                    bulan_index = 0  # Reset month index for the new year
                    
                    # Process monthly data
                    while bulan_index < len(bulan_list) and i < len(columns):
                        for sub in subcategories:
                            nama_kolom_baru.append(f"Claims {sub} {bulan_list[bulan_index]} {current_year}")
                            i += 1
                        
                        for sub in subcategories:
                            nama_kolom_baru.append(f"Liabilities {sub} {bulan_list[bulan_index]} {current_year}")
                            i += 1
                        bulan_index += 1
                
                elif columns[i] is not None and 1 < columns.count(columns[i]) <= 12 and columns[i] not in processed_years:
                    current_year = columns[i]
                    processed_years.append(current_year)
                    bulan_index = 0
                    
                    # Process monthly data
                    while bulan_index < len(bulan_list) and i < len(columns):
                        for sub in subcategories:
                            nama_kolom_baru.append(f"Claims {sub} {bulan_list[bulan_index]} {current_year}")
                            i += 1
                        
                        for sub in subcategories:
                            nama_kolom_baru.append(f"Liabilities {sub} {bulan_list[bulan_index]} {current_year}")
                            i += 1
                        bulan_index += 1
                
                else:
                    # Move to the next column if no valid conditions are met
                    i += 1
        
            df.columns = nama_kolom_baru
            data_frames.append(df)

        elif '5d' in token:  # Claims and liabilities
            subcategories = ['Total', 'Banks', 'Nonbanks']
            year_count = len([x for x in columns if x is not None])  # Hitung berapa banyak tahun di dalam kolom
            
            processed_years = []  # List to keep track of processed years

            while i < len(columns):
                if columns[i] is not None and columns.count(columns[i]) == 1 and columns[i] not in processed_years:
                    current_year = columns[i]
                    processed_years.append(current_year)
                    
                    # Loop through the subcategories for Claims
                    for sub in subcategories:
                        nama_kolom_baru.append(f"Claims {sub} {current_year}")
                        i += 1
                    
                    for sub in subcategories:
                        nama_kolom_baru.append(f"Liabilities {sub} {current_year}")
                        i += 1
                
                elif columns[i] is not None and columns.count(columns[i]) == 13 and columns[i] not in processed_years:
                    current_year = columns[i]
                    processed_years.append(current_year)
            
                    # Loop through the subcategories for Claims
                    for sub in subcategories:
                        nama_kolom_baru.append(f"Claims {sub} {current_year}")
                        i += 1
                    
                    for sub in subcategories:
                        nama_kolom_baru.append(f"Liabilities {sub} {current_year}")
                        i += 1
                    
                    bulan_index = 0  # Reset month index for the new year
                    
                    # Process monthly data
                    while bulan_index < len(bulan_list) and i < len(columns):
                        for sub in subcategories:
                            nama_kolom_baru.append(f"Claims {sub} {bulan_list[bulan_index]} {current_year}")
                            i += 1
                        
                        for sub in subcategories:
                            nama_kolom_baru.append(f"Liabilities {sub} {bulan_list[bulan_index]} {current_year}")
                            i += 1
                        bulan_index += 1
                
                elif columns[i] is not None and 1 < columns.count(columns[i]) <= 12 and columns[i] not in processed_years:
                    current_year = columns[i]
                    processed_years.append(current_year)
                    bulan_index = 0
                    
                    # Process monthly data
                    while bulan_index < len(bulan_list) and i < len(columns):
                        for sub in subcategories:
                            nama_kolom_baru.append(f"Claims {sub} {bulan_list[bulan_index]} {current_year}")
                            i += 1
                        
                        for sub in subcategories:
                            nama_kolom_baru.append(f"Liabilities {sub} {bulan_list[bulan_index]} {current_year}")
                            i += 1
                        bulan_index += 1
                
                else:
                    # Move to the next column if no valid conditions are met
                    i += 1
        
            df.columns = nama_kolom_baru
            data_frames.append(df)

        elif '5d1' in token:  # Claims and liabilities
            subcategories = ['Total', 'Related Offices']
            year_count = len([x for x in columns if x is not None])  # Hitung berapa banyak tahun di dalam kolom
            
            processed_years = []  # List to keep track of processed years

            while i < len(columns):
                if columns[i] is not None and columns.count(columns[i]) == 1 and columns[i] not in processed_years:
                    current_year = columns[i]
                    processed_years.append(current_year)
                    
                    # Loop through the subcategories for Claims
                    for sub in subcategories:
                        nama_kolom_baru.append(f"Claims {sub} {current_year}")
                        i += 1
                    
                    for sub in subcategories:
                        nama_kolom_baru.append(f"Liabilities {sub} {current_year}")
                        i += 1
                
                elif columns[i] is not None and columns.count(columns[i]) == 13 and columns[i] not in processed_years:
                    current_year = columns[i]
                    processed_years.append(current_year)
            
                    # Loop through the subcategories for Claims
                    for sub in subcategories:
                        nama_kolom_baru.append(f"Claims {sub} {current_year}")
                        i += 1
                    
                    for sub in subcategories:
                        nama_kolom_baru.append(f"Liabilities {sub} {current_year}")
                        i += 1
                    
                    bulan_index = 0  # Reset month index for the new year
                    
                    # Process monthly data
                    while bulan_index < len(bulan_list) and i < len(columns):
                        for sub in subcategories:
                            nama_kolom_baru.append(f"Claims {sub} {bulan_list[bulan_index]} {current_year}")
                            i += 1
                        
                        for sub in subcategories:
                            nama_kolom_baru.append(f"Liabilities {sub} {bulan_list[bulan_index]} {current_year}")
                            i += 1
                        bulan_index += 1
                
                elif columns[i] is not None and 1 < columns.count(columns[i]) <= 12 and columns[i] not in processed_years:
                    current_year = columns[i]
                    processed_years.append(current_year)
                    bulan_index = 0
                    
                    # Process monthly data
                    while bulan_index < len(bulan_list) and i < len(columns):
                        for sub in subcategories:
                            nama_kolom_baru.append(f"Claims {sub} {bulan_list[bulan_index]} {current_year}")
                            i += 1
                        
                        for sub in subcategories:
                            nama_kolom_baru.append(f"Liabilities {sub} {bulan_list[bulan_index]} {current_year}")
                            i += 1
                        bulan_index += 1
                
                else:
                    # Move to the next column if no valid conditions are met
                    i += 1
        
            df.columns = nama_kolom_baru
            data_frames.append(df)

        elif '5d2' in token:  # Claims and liabilities
            subcategories = ["Total", "Non-Bank Financial Ins.", "Non-Financial Corporations", "Others"]
            year_count = len([x for x in columns if x is not None])  # Hitung berapa banyak tahun di dalam kolom
            
            processed_years = []  # List to keep track of processed years

            while i < len(columns):
                if columns[i] is not None and columns.count(columns[i]) == 1 and columns[i] not in processed_years:
                    current_year = columns[i]
                    processed_years.append(current_year)
                    
                    # Loop through the subcategories for Claims
                    for sub in subcategories:
                        nama_kolom_baru.append(f"Claims {sub} {current_year}")
                        i += 1
                    
                    for sub in subcategories:
                        nama_kolom_baru.append(f"Liabilities {sub} {current_year}")
                        i += 1
                
                elif columns[i] is not None and columns.count(columns[i]) == 13 and columns[i] not in processed_years:
                    current_year = columns[i]
                    processed_years.append(current_year)
            
                    # Loop through the subcategories for Claims
                    for sub in subcategories:
                        nama_kolom_baru.append(f"Claims {sub} {current_year}")
                        i += 1
                    
                    for sub in subcategories:
                        nama_kolom_baru.append(f"Liabilities {sub} {current_year}")
                        i += 1
                    
                    bulan_index = 0  # Reset month index for the new year
                    
                    # Process monthly data
                    while bulan_index < len(bulan_list) and i < len(columns):
                        for sub in subcategories:
                            nama_kolom_baru.append(f"Claims {sub} {bulan_list[bulan_index]} {current_year}")
                            i += 1
                        
                        for sub in subcategories:
                            nama_kolom_baru.append(f"Liabilities {sub} {bulan_list[bulan_index]} {current_year}")
                            i += 1
                        bulan_index += 1
                
                elif columns[i] is not None and 1 < columns.count(columns[i]) <= 12 and columns[i] not in processed_years:
                    current_year = columns[i]
                    processed_years.append(current_year)
                    bulan_index = 0
                    
                    # Process monthly data
                    while bulan_index < len(bulan_list) and i < len(columns):
                        for sub in subcategories:
                            nama_kolom_baru.append(f"Claims {sub} {bulan_list[bulan_index]} {current_year}")
                            i += 1
                        
                        for sub in subcategories:
                            nama_kolom_baru.append(f"Liabilities {sub} {bulan_list[bulan_index]} {current_year}")
                            i += 1
                        bulan_index += 1
                
                else:
                    # Move to the next column if no valid conditions are met
                    i += 1
        
            df.columns = nama_kolom_baru
            data_frames.append(df)
    
    combined_df = pd.concat(data_frames, ignore_index=True)
    
    combined_df['Indikator'] = indikator_name  # Insert 'Indikator' at index 0
    combined_df['Path']= paths
    cols = combined_df.columns.tolist()
    new_order = cols[-2:] + cols[:-2]
    combined_df = combined_df[new_order]

    pd.set_option('display.max_rows', None)
    pd.set_option('display.max_columns', None)
    values_by_column = [combined_df.iloc[:, idx].tolist() for idx in range(1, combined_df.shape[1])]
    
    return values_by_column, combined_df

def gabungkan_dataframe_2bag(list_tahun, indikator_asli, indikator_calculated, mappings):
    # Mengolah data dengan fungsi dataframe_bag1
    values_bag1, df1 = dataframe_bag(indikator_asli, mappings)

    # Mengolah data dengan fungsi dataframe_bag2
    values_bag2, df2 = dataframe_bag(indikator_calculated, mappings)

    # Filter df1 based on list_tahun
    if list_tahun:
        columns_to_keep_bag1 = [df1.columns[0]] + [df1.columns[1]] + [col for col in df1.columns if any(tahun in col for tahun in list_tahun)]
    else:
        columns_to_keep_bag1 = df1.columns.tolist()

    df1_filtered = df1[columns_to_keep_bag1]

    # Filter df2 based on list_tahun
    if list_tahun:
        columns_to_keep_bag2 = [df2.columns[0]] + [df2.columns[1]] + [col for col in df2.columns if any(f'{tahun}' in col for tahun in list_tahun)]
    else:
        columns_to_keep_bag2 = df2.columns.tolist()
        
    df2_filtered = df2[columns_to_keep_bag2]

    common_columns = df1_filtered.columns.intersection(df2_filtered.columns).tolist()

    df1_filtered_common = df1_filtered[common_columns]
    df2_filtered_common = df2_filtered[common_columns]

    df_gabungan = pd.concat([df1_filtered_common, df2_filtered_common], ignore_index=True)
    df_gabungan = df_gabungan.round(2)

    periode = df_gabungan.columns.tolist()
    date_columns = periode[2:]

    if len(date_columns) == 1:
        periode_range = f"{date_columns[0]}"
    else:
        periode_range = f"{date_columns[0]} - {date_columns[-1]}"

    # Retrieve cleaned values for each source
    values_bag1_clean = [df1_filtered_common[col].tolist() for col in df1_filtered_common.columns]
    values_bag2_clean = [df2_filtered_common[col].tolist() for col in df2_filtered_common.columns]

    # Calculate jumlah_bag1
    jumlah_bag1 = []
    for sublist in values_bag1_clean[2:]:
        if "SSKI.6.RK.AJK.KPI" in indikator_asli or "SSKI.6.SOS.KPIAS" in indikator_asli:
            jumlah_bag1_value = 100
        else:
            jumlah_bag1_value = sum(sublist)
        jumlah_bag1.append(jumlah_bag1_value)

    # Calculate jumlah_bag2
    jumlah_bag2 = []
    for sublist in values_bag2_clean[2:]:
        if "-" in indikator_calculated:
            jumlah_bag2_value = sublist[0] - sum(sublist[1:])
        elif ";" in indikator_calculated:
            numeric_elements = [x for x in sublist if isinstance(x, (int, float))]
            jumlah_bag2_value = sum(numeric_elements) / len(numeric_elements) if numeric_elements else 0
        elif "/" in indikator_calculated:
            if len(sublist) > 1:
                jumlah_bag2_value = (sublist[0] / sublist[1]) * 100 if sublist[1] != 0 else 0
            else:
                jumlah_bag2_value = 0
        else:
            jumlah_bag2_value = sum(sublist)
        jumlah_bag2.append(jumlah_bag2_value)

    # Calculate selisih
    selisih = [round(abs(jumlah_bag1[i] - jumlah_bag2[i]), 2) for i in range(len(jumlah_bag1))]

    # Prepare 'selisih' row with 'Selisih' as the first entry
    selisih_row = [None] * (len(df_gabungan.columns))
    selisih_row[1] = 'Selisih'
    for idx, value in enumerate(selisih, start=2):
        selisih_row[idx] = value

    df_gabungan.loc[df_gabungan.shape[0]] = selisih_row

    value = df_gabungan.iloc[0, 0]

    # Create a DataFrame containing jumlah_bag1, jumlah_bag2, and selisih
    df_summary = pd.DataFrame({
        df_gabungan.columns[0]: [value, value + ' (Calculated)', 'Selisih'],
    })
    
    for i, col in enumerate(df_gabungan.columns[2:], start=2):
        df_summary[col] = [jumlah_bag1[i-2], jumlah_bag2[i-2], selisih[i-2]]

    df_raw = df_gabungan.copy()
    # Drop columns where the "Selisih" row has a value of 0
    columns_to_drop = df_gabungan.columns[df_gabungan.iloc[-1] == 0]
    df_gabungan.drop(columns=columns_to_drop, inplace=True)

    # Copy df_gabungan before column removal (raw version)

    return df_gabungan, df_raw, df_summary

# HORIZONTAL CHECK
# HORIZONTAL CHECK
# HORIZONTAL CHECK

super_sheet = pd.ExcelFile("/Users/ferroyudisthira/Desktop/DSTA_DQAD/V&H_Check/Sumber_Data_Lama/SSKI/SSKI EKSTERNAL_September 2024 v3.xlsx")
import datetime
year_to_check = datetime.date.today().year - 1
globals()[str(year_to_check)] = year_to_check

def prepare_dataframe(sheet_name, header_count):
    input_df = super_sheet.parse(sheet_name, header=[4, 3+header_count])

    input_df.columns = input_df.columns.map(lambda x: str(x[0]).upper() if 'Unnamed' in str(x) else x)

    input_df.columns = input_df.columns.map(lambda x: (x[0], x[1].strip()) if isinstance(x, tuple) and isinstance(x[1], str) else x)

    first_column = input_df.iloc[:, 0]
    second_column = input_df.iloc[:, 1]

    output_df = input_df.filter(regex=str(year_to_check))

    output_df.insert(0, 'NO', first_column)
    output_df.insert(1, 'Komponen', second_column)

    index_keterangan = output_df[output_df.apply(lambda row: row.astype(str).str.contains('keterangan', case=False).any(), axis=1)].index

    if not index_keterangan.empty:
        first_keterangan_index = index_keterangan[0]
        output_df = output_df.iloc[:first_keterangan_index]

    return(output_df)

def prepare_dataframe_5(sheet_name): 
    if sheet_name == '5a':
        header_array=[3,4,5]
    else:
        header_array=[2,3,4,5]

    input_df = super_sheet.parse(sheet_name, header=header_array)
    
    input_df.columns = input_df.columns.map(
        lambda x: tuple(
            str(x[i]).upper() if i == 0 and 'Unnamed' in str(x[1]) else x[i].strip() if isinstance(x[i], str) else x[i]
            for i in range(len(x))
        ) if isinstance(x, tuple) else x
    )
    # make sure to include all column within the current year to be checked
    output_df = input_df.filter(regex=str(year_to_check))
    output_df.insert(0, 'NO', input_df[list(input_df.columns)[0]])
    output_df.insert(1, 'Komponen', input_df[list(input_df.columns)[1]])
    
    # Remove row with "keterangan" and rows following
    index_keterangan = output_df[output_df.apply(lambda row: row.astype(str).str.contains('keterangan', case=False).any(), axis=1)].index

    if not index_keterangan.empty:
        first_keterangan_index = index_keterangan[0]
        output_df = output_df.iloc[:first_keterangan_index]

    return(output_df)

def masking(df, key, value):
    return df[df[key] == value]

def average_formula(input_df,index):
    to_check = masking(input_df, "NO", index)
    to_check = to_check.iloc[:,3:]
    to_check = to_check.drop(to_check.columns[-1], axis=1)
    to_check_avg =np.ceil(float(to_check.mean(axis=1)))
    data_real = np.ceil(float(input_df.loc[input_df["NO"] == index, str(year_to_check)]))

    if to_check_avg == data_real:
        input_df = input_df.drop(index-1)
    
    return input_df

def sum_formula(input_df,index):
    to_check = masking(input_df, "NO", index)
    to_check = to_check.iloc[:,3:]
    to_check = to_check.drop(to_check.columns[-1], axis=1)
    to_check_avg =np.ceil(float(to_check.sum(axis=1)))
    data_real = np.ceil(float(input_df.loc[input_df["NO"] == index, str(year_to_check)]))

    if to_check_avg == data_real:
        input_df = input_df.drop(index-1)
    
    return input_df

def h1_score_year_month(input_df, sheet_name):
    year_str = str(year_to_check)

    # Get two arrays for comparison
    arr1 = pd.to_numeric(input_df.filter(like=year_str).iloc[:, 0], errors='coerce').to_numpy()

    # Initialize arr2
    arr2 = np.zeros_like(arr1)  # Default arr2 to zeros if not overridden

    # Check for Dec column or fallback to Q4
    if sheet_name == "13":
        # Select the relevant columns
        to_check = input_df.iloc[:, 3:]

        # Replace '-' with NaN, then convert to numeric
        to_check.replace('-', np.nan, inplace=True)
        to_check = to_check.apply(pd.to_numeric, errors='coerce').fillna(0)

        # Calculate the mean and round it up
        mean_values = np.ceil(to_check.mean(axis=1))

        # Convert the mean values to a numpy array
        arr2 = mean_values.to_numpy()
    else:    
        if (year_to_check, 'Dec') in input_df.columns:
            arr2 = pd.to_numeric(input_df[(year_to_check, 'Dec')], errors='coerce').to_numpy()
        elif (year_to_check, 'Des') in input_df.columns:
            arr2 = pd.to_numeric(input_df[(year_to_check, 'Des')], errors='coerce').to_numpy()
        else:
            arr2 = pd.to_numeric(input_df.filter(like='Q4').iloc[:, 0], errors='coerce').to_numpy()

    # Mask for valid and non-NaN values
    valid_mask = ~np.isnan(arr1) & ~np.isnan(arr2)

    # Find mismatches where arr1 != arr2
    mismatches = valid_mask & (arr1 != arr2)

    # Calculate differences for mismatches
    differences = np.where(mismatches, np.abs(arr1 - arr2), 0)

    # Add the H1 Score column with the differences (no filtering)
    input_df['H1 SCORE'] = differences

    # HARDCODED HARDCODED HARDCODED
    if sheet_name == "1":
        input_df = average_formula(input_df, 34)
        input_df = sum_formula(input_df, 35)

        ignore_array = [2, 5, 28, 29, 30]
        for item in ignore_array:
            input_df.loc[item - 1, 'H1 SCORE'] = "TBC"
    if sheet_name == "2":
        input_df.replace('-', 0, inplace=True)
        sum_array = [39, 40, 42, 65, 66, 67, 68, 69, 70, 71, 72, 73, 74, 75, 76, 77, 78, 79, 80, 81, 82]
        for item in sum_array:
            input_df = sum_formula(input_df, item)
    if sheet_name == "13":
        input_df.replace('-', 0, inplace=True)
        sum_array = [37, 39, 40, 41]
        for item in sum_array:
            input_df = sum_formula(input_df, item)
    if sheet_name == "14":
        input_df.replace('-', 0, inplace=True)

        average_array = [31, 33, 38, 42, 43, 44, 45, 46, 47]
        for item in average_array:
            input_df = average_formula(input_df, item)
        
        sum_array = [35, 36, 37]
        for item in sum_array:
            input_df = sum_formula(input_df, item)

        ignore_array = [32, 34, 39]
        for item in ignore_array:
            input_df.loc[item - 1, 'H1 SCORE'] = "TBC"

    # Create a DataFrame containing only the rows with mismatches
    mismatch_df = input_df.loc[input_df['H1 SCORE'] != 0] 

    # Return the entire dataframe and the mismatch dataframe
    return input_df, mismatch_df

def h1_score_claim_liabilities(input_df, sheet_name):
    if sheet_name != "5a":
        last_level_columns = input_df.columns.get_level_values(-1)
        unique_subcategories = np.array(last_level_columns.unique())[1:]  
    
    middle_level_columns = input_df.columns.get_level_values(2)
    unique_categories = np.array(middle_level_columns.unique())[1:]

    # Loop through all the unique categories (Claims, Liabilities) and their subcategories
    for cat in unique_categories:
        if sheet_name == "5a":
            # Select the first and last columns matching the category
            arr1 = pd.to_numeric(input_df.filter(like=f'{cat}').iloc[:, 0], errors='coerce').fillna(0).to_numpy()
            arr2 = pd.to_numeric(input_df.filter(like=f'{cat}').iloc[:, -1], errors='coerce').fillna(0).to_numpy()

            # Calculate mismatches where the values are different
            mismatches = (arr1 != arr2)

            # Calculate the absolute difference where mismatches occur, otherwise fill with 0
            differences = np.where(mismatches, np.abs(arr1 - arr2), 0)

            # Assign the differences to a new column in the dataframe
            input_df[f'{cat} H1 SCORE'] = differences
        else:
            for item in unique_subcategories:
                # Select the first and last columns matching the category and subcategory
                arr1 = pd.to_numeric(input_df.filter(like=f'{cat}').filter(like=item).iloc[:, 0], errors='coerce').fillna(0).to_numpy()
                arr2 = pd.to_numeric(input_df.filter(like=f'{cat}').filter(like=item).iloc[:, -1], errors='coerce').fillna(0).to_numpy()

                # Calculate mismatches where the values are different
                mismatches = (arr1 != arr2)

                # Calculate the absolute difference where mismatches occur, otherwise fill with 0
                differences = np.where(mismatches, np.abs(arr1 - arr2), 0)

                # Assign the differences to a new column in the dataframe
                input_df[f'{cat} {item} H1 SCORE'] = differences

    # Use filter(regex=...) to capture all columns with 'H1 SCORE' in their names
    score_columns = input_df.filter(regex='H1 SCORE')

    # Filter the input_df where any of the H1 SCORE columns have non-zero values
    mismatch_df = input_df[(score_columns != 0).any(axis=1)]

    ## Flatten the multi-level columns to a single level
    input_df.columns = [' '.join([str(i) for i in col if 'Unnamed' not in str(i)]).strip() if isinstance(col, tuple) else str(col) for col in input_df.columns]
    mismatch_df.columns = [' '.join([str(i) for i in col if 'Unnamed' not in str(i)]).strip() if isinstance(col, tuple) else str(col) for col in mismatch_df.columns]
    return input_df, mismatch_df

def run_horizontal_check(sheet_to_check, header_column):
    if sheet_to_check.startswith("5"):
        start_df = prepare_dataframe_5(sheet_to_check)
        final, clean = h1_score_claim_liabilities(start_df, sheet_to_check)
    else:
        start_df = prepare_dataframe(sheet_to_check, header_column)
        final, clean = h1_score_year_month(start_df, sheet_to_check)
    
    # Ensure the output folder exists, create it if necessary
    FILEPATH = os.path.join(os.getcwd(), "SAVED")
    if not os.path.exists(FILEPATH):
        os.makedirs(FILEPATH)

    output_filename = os.path.join(FILEPATH, f'SSKI TABLE {sheet_to_check} LOG.xlsx')
    
    # Check if the Excel file exists
    if os.path.exists(output_filename):
        wb = openpyxl.load_workbook(output_filename)

        if 'Horizontal Check' in wb.sheetnames:
            sheet = wb['Horizontal Check']
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
                for cell in row:
                    cell.value = None
        else:
            sheet = wb.create_sheet('Horizontal Check')
    else:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = 'Horizontal Check'

    # Write the header to the "Horizontal Check" sheet
    for c_idx, column_name in enumerate(final.columns, 1):
        # Convert tuple column names to string if necessary
        if isinstance(column_name, tuple):
            column_name = ' '.join(map(str, column_name))  # Convert to space-separated string
        else:
            column_name = str(column_name)  # Ensure it's a string if it's not a tuple
        sheet.cell(row=1, column=c_idx, value=column_name)

    # Function to safely convert values to string
    def convert_value(value):
        if isinstance(value, (tuple, list)):
            return ', '.join(map(str, value))
        elif pd.isna(value):
            return ''
        else:
            return str(value)

    # Write the DataFrame rows (excluding header) to the sheet
    for r_idx, row in enumerate(final.itertuples(index=False), 2):
        for c_idx, value in enumerate(row, 1):
            sheet.cell(row=r_idx, column=c_idx, value=convert_value(value))

    # Define the thin border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Set a default font for the header and data
    header_font = Font(bold=True)
    data_font = Font()
    center_alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

    # Apply styles to the header row
    for cell in sheet[1]:
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border

    # Apply styles to the data rows
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        for cell in row:
            cell.font = data_font
            cell.alignment = center_alignment
            cell.border = thin_border

    # Define the column widths
    column_widths = {
        'A': 5, 'B': 60
    }

    # Apply column widths
    for col, width in column_widths.items():
        sheet.column_dimensions[col].width = width

    # Save the workbook with the updated data
    wb.save(output_filename)

    return final, clean, f"Data has been written to {output_filename}"


# Set Streamlit to use the wider layout mode
st.set_page_config(layout="wide", page_title="Kelompok Data Viewer")

def set_border(sheet, cell_range):
    thin = Side(border_style="thin", color="000000")
    for row in sheet[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

def save_ver(df_raw):
    # Extract the number after 'SSKI.' from the "Path" column
    extracted_numbers = df_raw['Path'].str.extract(r'SSKI\.([^\.]+)\.[^\.]*')[0]

    # Set up the file path
    if os.path.exists("./SAVED"):
        FILEPATH = os.getcwd() + "/SAVED"
    else:
        FILEPATH = os.path.join(os.getcwd(), "SAVED")
        os.makedirs(FILEPATH, exist_ok=True)  # Create the directory if it doesn't exist

    # Extract the number for the filename
    num = extracted_numbers.dropna().tolist()[0] if not extracted_numbers.empty else None
    if num == '5d1':
        num = '5d.1'
    elif num == '5d2':
        num = '5.d.2'

    # File path for saving
    output_filename = os.path.join(FILEPATH, f'SSKI TABLE {num} LOG.xlsx')

    # Load the existing workbook or create a new one
    if os.path.exists(output_filename):
        wb = load_workbook(output_filename)
        # Create 'VERTICAL CHECK' sheet if it doesn't exist
        if 'Vertical Check' in wb.sheetnames:
            sheet = wb['Vertical Check']
        else:
            sheet = wb.create_sheet(title='Vertical Check')
    else:
        wb = Workbook()
        sheet = wb.active
        sheet.title = 'Vertical Check'

    # Find the next empty row
    next_row = sheet.max_row + 2  # This will point to the row after the last filled row

    # Write the headers for the current DataFrame
    for c_idx, value in enumerate(df_raw.columns, start=1):
        sheet.cell(row=next_row, column=c_idx, value=value)

    # Set the width for the first and second columns only and wrap the text
    for col_idx in [1, 2]:  # Only first two columns
        col_letter = get_column_letter(col_idx)
        sheet.column_dimensions[col_letter].width = 40
        for row in sheet.iter_rows(min_row=next_row, max_row=sheet.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)

    # Move to the next row for data
    next_row += 1

    # Append the new DataFrame below the existing data
    for r_idx, row in df_raw.iterrows():
        for c_idx, value in enumerate(row, start=1):  # Start from column 1 (A)
            cell = sheet.cell(row=next_row, column=c_idx, value=value)
            if c_idx in [1, 2]:  # Apply wrapping to the first two columns
                cell.alignment = Alignment(wrap_text=True)
        next_row += 1  # Move to the next row after writing each row of the DataFrame

    # Apply borders to the table (from headers to last row)
    total_rows = len(df_raw) + 1  # +1 to include header
    cell_range = f"A{next_row - total_rows}:{get_column_letter(len(df_raw.columns))}{next_row - 1}"
    set_border(sheet, cell_range)
    wb.save(output_filename)

def main(list_tahun, indikator_asli, indikator_calculated, mappings, kelompok):
    data_clean_array = {}
    data_raw_array = {}
    data_summary_array = {}
    data_horizontal_array = {}
    data_raw_hor_array = {}

    for i in range(len(indikator_asli)):
        # Combine indicators and skip if invalid
        df_clean, df_raw, df_summary = gabungkan_dataframe_2bag(list_tahun, indikator_asli[i], indikator_calculated[i], mappings)
        
        sski_path = df_clean['Path'][0]  # Assuming 'Path' column exists
        sski_number = sski_path.split('.')[1]

        json_raw_df = df_raw.to_json(orient='records')
        data_raw_array[f"{sski_number}-{kelompok[i]}"] = json.loads(json_raw_df)

        json_clean_df = df_clean.to_json(orient='records')
        data_clean_array[f"{sski_number}-{kelompok[i]}"] = json.loads(json_clean_df)

        json_summary_df = df_summary.to_json(orient='records')
        data_summary_array[f"{sski_number}-{kelompok[i]}"] = json.loads(json_summary_df)
    
    # Horizontal check table list and error counts
    table_list = ["1", "2", "3", "4", "5a", "5b", "5c", "5d", "5d.1", "5.d.2", "6", "7", "8", "9", "10", "11a", "12", "13", "14", "15", "16a", "17", "18", "19", "20"]
    horizontal_results = {}

    # Assuming you still want to run horizontal checks
    for item in table_list:
        final, clean, msg = run_horizontal_check(item, 2)
        if clean is not None and not clean.empty:
            json_horizontal_df = clean.to_json(orient='records')
            data_horizontal_array[item] = json.loads(json_horizontal_df)  # Save clean DataFrame for later display
        json_horizontal_raw_df = final.to_json(orient='records')
        data_raw_hor_array[item] = json.loads(json_horizontal_raw_df)
        
            
    combined_data = {
        "clean_data": data_clean_array,
        "raw_data": data_raw_array,
        "summary_data": data_summary_array,
        "horizontal_clean_data": data_horizontal_array,
        "horizontal_raw_data": data_raw_hor_array
    }

    # Save the combined JSON to a file
    output_file_path = '/Users/ferroyudisthira/Desktop/DSTA_DQAD/V&H_Check/data_test.json'
    with open(output_file_path, 'w') as json_file:
        json.dump(combined_data, json_file, indent=4)

    print(f"Data has been written to {output_file_path}")

# Example parameters (replace with your actual data)
list_tahun = ['2022']

# Assuming indikator_asli, indikator_calculated, mappings, and kelompok are defined
main(list_tahun, indikator_asli, indikator_calculated, mappings, kelompok)
